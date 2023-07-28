import os
from openpyxl import load_workbook, Workbook
from labels_mac import make_labels


def new_book():
    shk_excel = Workbook()
    work_sheet = shk_excel.active
    work_sheet.title = 'Sheet1'
    barcode_title = 'баркод товара'
    amount_title = 'кол-во товаров'
    boxes_title = 'шк короба'
    best_before = 'срок годности'

    work_sheet.cell(row=1, column=1, value=barcode_title)
    work_sheet.cell(row=1, column=2, value=amount_title)
    work_sheet.cell(row=1, column=3, value=boxes_title)
    work_sheet.cell(row=1, column=4, value=best_before)

    return {'book': shk_excel,
            'sheet': work_sheet}


def shipment_book(shipment_total, shipment_filename):
    shipment = Workbook()
    work_sheet = shipment.active
    work_sheet.title = 'Sheet1'
    barcode_title = 'Баркод'
    amount_title = 'Количество'
    barcode_format = '0'

    work_sheet.cell(row=1, column=1, value=barcode_title)
    work_sheet.cell(row=1, column=2, value=amount_title)
    
    row_count = 2

    for bc, amount in shipment_total.items():
        work_sheet.cell(row=row_count,
                        column=1,
                        value=bc).number_format = barcode_format
        
        work_sheet.cell(row=row_count,
                        column=2,
                        value=amount)
        row_count += 1

    shipment.save('output_new' + '//' + shipment_filename)


def iterate_boxes(boxes, wb_array, dictionary,
                  wb_barcodes_filename, shipment_filename):
    book_and_sheet = new_book()
    shk_excel = book_and_sheet.get('book')
    work_sheet = book_and_sheet.get('sheet')
    barcode_format = '0'
    row_count = 2
    shipment_total = {}
    step = -1
    box_id = wb_array[step]
    
    for row in boxes.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            box_id = box_id
        else:
            step += 1
            box_id = wb_array[step]
        item_id = row[1]
        color = row[2]
        size = row[3]
        amount = row[4]
        barcode = int(dictionary[str(item_id)][str(size)][color])
        
        # print(barcode, amount, box_id)
        
        work_sheet.cell(row=row_count,
                        column=1, 
                        value=barcode).number_format = barcode_format
        work_sheet.cell(row=row_count, column=2, value=amount)
        work_sheet.cell(row=row_count, column=3, value=box_id)
        row_count += 1
        
        if barcode in shipment_total.keys():
            total_amount = shipment_total.get(barcode) + amount
            shipment_total.update({barcode: total_amount})
        else:
            shipment_total.update({barcode: amount})

    shipment_book(shipment_total, shipment_filename)
    shk_excel.save('output_new' + '//' + wb_barcodes_filename)


def iterate_wb_boxes(wb_boxes):
    wb_array = []

    for row in wb_boxes.iter_rows(
        min_row=2, 
        min_col=3, 
        max_col=3,
        values_only=True
    ):
        wb_array.append(row[0])

    return wb_array


def excel(dictionary, plan_file, wb_boxes_file,
          wb_barcodes_filename, shipment_filename):
    plan = load_workbook(filename=plan_file)
    wb_boxes = load_workbook(filename=wb_boxes_file).active
    
    boxes = plan['КОРОБА']
    wb_array = iterate_wb_boxes(wb_boxes)
    iterate_boxes(boxes, wb_array, dictionary,
                  wb_barcodes_filename, shipment_filename)

  
def bar_codes(barcodes_file):
    dictionary = {}
    work_book = load_workbook(filename=barcodes_file)
    table = work_book['Лист1']
    exclude = 'Артикул поставщика'

    count = 0
    for row in table.iter_rows(values_only=True):
        if row[0] != exclude and row[0] is not None:
            count += 1
            item_id = row[0][5:]
            size = row[1][7:]
            color = row[2][8:]
            barcode = row[3]

            if item_id not in dictionary.keys():
                dictionary.update({item_id: {}})
                dictionary[item_id].update({size: {}})
                dictionary[item_id][size].update({color: barcode})
            elif size not in dictionary[item_id].keys():
                dictionary[item_id].update({size: {}})
                dictionary[item_id][size].update({color: barcode})
            elif color not in dictionary[item_id][size].keys():
                dictionary[item_id][size].update({color: barcode})
    return dictionary
    

def get_files():
    directory = os.listdir()
    plan = 'План поставки'
    wb = 'shk-excel'
    barcodes = 'ШК 20'

    plan_file = ''
    wb_boxes_file = ''
    barcodes_file = ''
    
    for i in directory:
        if plan in i:
            plan_file = i
        elif wb in i:
            wb_boxes_file = i
        elif barcodes in i:
            barcodes_file = i

    return {'plan': plan_file,
            'wb': wb_boxes_file,
            'barcodes': barcodes_file} 


def main():
    shipment_date = make_labels()
    files = get_files()
    
    wb_barcodes_filename = 'shk_excel-' + shipment_date + '.xlsx'
    shipment_filename = 'Поставка ' + shipment_date + '.xlsx'
    
    plan_file = files.get('plan')
    wb_boxes_file = files.get('wb')
    barcodes_file = files.get('barcodes')

    if (plan_file is not None and
        wb_boxes_file is not None and
        barcodes_file is not None):

        dictionary = bar_codes(barcodes_file)
        excel(dictionary, plan_file,
              wb_boxes_file, wb_barcodes_filename,
              shipment_filename)

    input('Всё готово и сложено в папку "output_new"')


if __name__ == '__main__':
    main()
