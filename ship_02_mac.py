import os, sys, re
from openpyxl import load_workbook, Workbook
from labels import make_labels


def new_book():
    shk_excel = Workbook()
    work_sheet = shk_excel.active
    work_sheet.title = 'Sheet1'
    barcode_title = 'баркод товара'
    amount_title = 'кол-во товаров'
    boxes_title = 'шк короба'
    best_before = 'срок годности'

    work_sheet.cell(row=1, column=1, value=barcode_title)
    work_sheet.cell(row=1, column=2, value=amount_title)
    work_sheet.cell(row=1, column=3, value=boxes_title)
    work_sheet.cell(row=1, column=4, value=best_before)

    return {'book': shk_excel,
            'sheet': work_sheet}


def shipment_book(shipment_total, shipment_filename):
    shipment = Workbook()
    work_sheet = shipment.active
    work_sheet.title = 'Sheet1'
    barcode_title = 'Баркод'
    amount_title = 'Количество'
    barcode_format = '0'

    work_sheet.cell(row=1, column=1, value=barcode_title)
    work_sheet.cell(row=1, column=2, value=amount_title)
    
    row_count = 2

    for bc, amount in shipment_total.items():
        work_sheet.cell(row=row_count,
                        column=1,
                        value=bc).number_format = barcode_format
        
        work_sheet.cell(row=row_count,
                        column=2,
                        value=amount)
        row_count += 1

    shipment.save('output_new' + '//' + shipment_filename)


def iterate_boxes(boxes,
                  dictionary,
                  wb_barcodes_filename,
                  shipment_filename): 
    book_and_sheet = new_book()
    shk_excel = book_and_sheet.get('book')
    work_sheet = book_and_sheet.get('sheet')
    barcode_format = '0'
    row_count = 2
    shipment_total = {}
    box_id = 0
    
    for row in boxes.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            box_id += 1

        item_id = row[1]
        color = row[2]
        size = row[3]
        amount = row[4]
        barcode = int(dictionary[str(item_id)][str(size)][color])
        
        work_sheet.cell(row=row_count,
                        column=1, 
                        value=barcode).number_format = barcode_format
        work_sheet.cell(row=row_count, column=2, value=amount)
        work_sheet.cell(row=row_count, column=3, value=box_id)
        row_count += 1
        
        if barcode in shipment_total.keys():
            total_amount = shipment_total.get(barcode) + amount
            shipment_total.update({barcode: total_amount})
        else:
            shipment_total.update({barcode: amount})

    shipment_book(shipment_total, shipment_filename)
    shk_excel.save('output_new' + '//' + wb_barcodes_filename)


def iterate_wb_boxes(wb_boxes):
    wb_array = []

    for row in wb_boxes.iter_rows(
        min_row=2, 
        min_col=3, 
        max_col=3,
        values_only=True
    ):
        wb_array.append(row[0])

    return wb_array


def excel(dictionary, plan_file,
          wb_barcodes_filename, shipment_filename):
    plan = load_workbook(filename=plan_file)
    
    boxes = plan['КОРОБА']
    iterate_boxes(boxes, 
                  dictionary,
                  wb_barcodes_filename, shipment_filename)

  
def bar_codes(barcodes_file):
    dictionary = {}
    work_book = load_workbook(filename=barcodes_file)
    table = work_book['Лист1']
    exclude = 'Артикул поставщика'

    count = 0
    for row in table.iter_rows(values_only=True):
        if row[0] != exclude and row[0] is not None:
            count += 1
            item_id = row[0][5:]
            size = row[1][7:]
            color = row[2][8:]
            barcode = row[3]

            if item_id not in dictionary.keys():
                dictionary.update({item_id: {}})
                dictionary[item_id].update({size: {}})
                dictionary[item_id][size].update({color: barcode})
            elif size not in dictionary[item_id].keys():
                dictionary[item_id].update({size: {}})
                dictionary[item_id][size].update({color: barcode})
            elif color not in dictionary[item_id][size].keys():
                dictionary[item_id][size].update({color: barcode})
    return dictionary
    

def get_files():
    directory = os.listdir()
    plan = 'План поставки'
    wb = 'shk-excel'
    barcodes = 'ШК 20'

    plan_file = ''
    wb_boxes_file = ''
    barcodes_file = ''
    
    for i in directory:
        if plan in i:
            plan_file = i
        elif barcodes in i:
            barcodes_file = i
        elif wb in i:
            wb_boxes_file = i
        

    return {'plan': plan_file,
            'barcodes': barcodes_file,
            'wb': wb_boxes_file}


def check_if_ready(files, wb_barcodes_filename):
    if 'output_new' in os.listdir():
        os.chdir('output_new')
        output_dir = os.listdir()

        for file in output_dir:
            if 'shk-excel' in file:
                print('Уже всё готово, можно отправлять!')
                sys.exit()
        
        if wb_barcodes_filename in output_dir:
            wb_boxes_file = files.get('wb')
            wb_boxes = load_workbook(filename='..//' + wb_boxes_file).active
            wb_array = iterate_wb_boxes(wb_boxes)

            wb_shipment_file = load_workbook(filename=wb_barcodes_filename)
            wb_shipment_file_sheet = wb_shipment_file.active
            wb_id = 1

            for row in wb_shipment_file_sheet.iter_rows(min_row=2):
                if row[2].value >= wb_id:
                    wb_id = row[2].value
                    row[2].value = wb_array[wb_id - 1]
        
            new_book_name = 'shk-excel-NEW.xlsx'
            for filename in os.listdir():
                if 'Поставка' in filename:
                    find_index = re.search('\\d', filename).start()
                    current_date = filename[find_index:]
                    new_book_name = 'shk-excel-' + current_date + '.xlsx'

            wb_shipment_file.save(new_book_name)
            os.remove(wb_barcodes_filename)
            input('Всё готово и сложено в папку \'output_new\'')
            sys.exit()

        os.chdir('..')


def main():
    files = get_files()
    wb_barcodes_filename = 'pre_shipped_wb_shk.xlsx'
    check_if_ready(files, wb_barcodes_filename)
    shipment_date = make_labels()
    shipment_filename = 'Поставка ' + shipment_date + '.xlsx'
    plan_file = files.get('plan')
    barcodes_file = files.get('barcodes')

    if (plan_file is not None and
        barcodes_file is not None):
    
        dictionary = bar_codes(barcodes_file)
        excel(dictionary, plan_file,
              wb_barcodes_filename,
              shipment_filename)

    input('Файл поставки готов. Пора его загружать и скачать файл с кодами WB.')


if __name__ == '__main__':
    main()
